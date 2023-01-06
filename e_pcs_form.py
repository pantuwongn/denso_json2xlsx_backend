from openpyxl import load_workbook
from openpyxl.drawing.xdr import XDRPositiveSize2D, XDRPoint2D
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker, AbsoluteAnchor
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.drawing.image import Image
from openpyxl.drawing.text import TextField
from openpyxl.utils.units import cm_to_EMU, pixels_to_EMU, EMU_to_pixels
from utils import (
    getOutputFilePath,
    drawVerticalDashedLine,
    chunk,
    leftCenterAlignment,
    centerCenterAlignment,
    topCenterAlignment,
    topLeftAlignment,
    headerNormalStyle,
    textNormalStyle,
    bottomBorder,
    bottomRightBorder
)

itemChunkSize = 17

timingConnectorPath = 'images/timing/check-process.png'

counterPathMap = {
    1: 'images/counter/1.png',
    2: 'images/counter/2.png',
    3: 'images/counter/3.png',
    4: 'images/counter/4.png',
    5: 'images/counter/5.png',
    6: 'images/counter/6.png',
    7: 'images/counter/7.png',
    8: 'images/counter/8.png',
    9: 'images/counter/9.png',
    10: 'images/counter/10.png',
    11: 'images/counter/11.png',
    12: 'images/counter/12.png',
    13: 'images/counter/13.png',
    14: 'images/counter/14.png',
    15: 'images/counter/15.png',
    16: 'images/counter/16.png',
}

checkTimingSymbolPathMap = {
    'None': 'images/timing/check-no-record.png',
    'Check sheet': 'images/timing/check-record.png',
    'Record sheet': 'images/timing/check-record.png',
    'x-R chart': 'images/timing/check-control-chart.png',
    'xbar-R chart': 'images/timing/check-control-chart.png',
    'x-Rs chart': 'images/timing/check-control-chart.png',
}

scSymbolPathMap = {
    'C-none': 'images/symbols/C-none.png',
    'S-circle': 'images/symbols/S-circle.png',
    'S-diamond': 'images/symbols/S-diamond.png',
    'F-circle': 'images/symbols/F-circle.png',
    'F-triangle': 'images/symbols/F-triangle.png',
    'RW-rectangle': 'images/symbols/RW-rectangle.png',
    'SP-circle': 'images/symbols/SP-circle.png'
}

c2e = cm_to_EMU
p2e = pixels_to_EMU
cellh = lambda x: c2e(x * 0.48)
cellw = lambda x: c2e(x * 1.1)

def getSCSymbolList(scSymbolList: list, rowStart: int, maxRow: int):
    imgList = list()
    symbolTotal = len(scSymbolList)

    assert symbolTotal <= maxRow, 'SC Symbol out of bound'

    for i, scSymbol in enumerate(scSymbolList):
        symbolPath = scSymbolPathMap.get('{}-{}'.format(scSymbol['character'], scSymbol['shape']), None)
        if symbolPath is None:
            raise KeyError('Unregistered sc symbol, {}-{}'.format(scSymbol['character'], scSymbol['shape']))

        if symbolTotal == 1:
            symbolImg = drawImage(Image(symbolPath), rowStart, 2, 0, 10)
        else:
            symbolImg = drawImage(Image(symbolPath), rowStart + i -1, 2, 5*i, 10)
        imgList.append(symbolImg)
    return imgList

def getTotalSCSymbolList(itemList: list):
    def drawTotalScSymbol(img, rowOff, colOff):
        h, w = img.height, img.width
        size = XDRPositiveSize2D(p2e(w), p2e(h))
        position = XDRPoint2D(p2e(520 + colOff), p2e(135 + rowOff))
        img.anchor = AbsoluteAnchor(pos=position, ext=size)
        return img
    def drawTotalCountSymbol(img, rowOff, colOff):
        h, w = img.height, img.width
        size = XDRPositiveSize2D(p2e(w), p2e(h))
        position = XDRPoint2D(p2e(520 + colOff), p2e(135 + rowOff))
        img.anchor = AbsoluteAnchor(pos=position, ext=size)
        return img
    scSymbolDict = dict()
    scSymbolCountDict = dict()
    for item in itemList:
        scSymbolList = item['sc_symbols']
        for scSymbol in scSymbolList:
            symbolHash = '{}-{}'.format(scSymbol['character'], scSymbol['shape'])
            scSymbolDict[symbolHash] = scSymbol

            if scSymbolCountDict.get(symbolHash, None) is None:
                scSymbolCountDict[symbolHash] = 1
            else:
                scSymbolCountDict[symbolHash] += 1

    totalSymbolList = list(scSymbolDict.values())

    imgList = list()
    for i, scSymbol in enumerate(totalSymbolList):
        symbolHash = '{}-{}'.format(scSymbol['character'], scSymbol['shape'])
        symbolPath = scSymbolPathMap.get(symbolHash, None)

        if symbolPath is None:
            raise KeyError('Unregistered sc symbol, {}-{}'.format(scSymbol['character'], scSymbol['shape']))

        symbolImg = drawTotalScSymbol(Image(symbolPath), 0, 33*i)
        counterImg = drawTotalCountSymbol(Image(counterPathMap[scSymbolCountDict[symbolHash]]), 12, (33 * i) + 23)
        imgList.append(symbolImg)
        imgList.append(counterImg)
    
    return imgList

def getProcessCapability(capabilityDict: dict):
    x_bar = 'xbar : {}'.format(capabilityDict['x_bar']) if capabilityDict['x_bar'].strip() != '' else ''
    cpk = 'cpk : {}'.format(capabilityDict['cpk']) if capabilityDict['cpk'].strip() != '' else ''

    result = ''
    if x_bar != '':
        result = '{}{}\n'.format(result, x_bar)
    if cpk != '':
        result = '{}{}'.format(result, cpk)

    return result

def getParameter(parameterDict: dict):
    limitType = parameterDict['limit_type']
    if limitType == 'None':
        return parameterDict['parameter']


    def _appendTextIfExist(targetStr: str, dataDict: dict):
        def doAppendTextIfExist(key: str):
            finalText = targetStr
            if dataDict.get(key, '').strip() != '':
                finalText = '{}{}'.format(targetStr, dataDict[key])
            return finalText
        return doAppendTextIfExist

    result = '{}\n'.format(parameterDict['parameter'])
    appendTextIfExist = _appendTextIfExist(result, parameterDict)
    result = appendTextIfExist('prefix')
    result = appendTextIfExist('main')
    result = appendTextIfExist('suffix')
    result = appendTextIfExist('tolerance_up')
    result = appendTextIfExist('tolerance_down')
    result = appendTextIfExist('unit')

    return result

def getInterval(controlMethodDict: dict):
    intervalText = controlMethodDict['interval']
    if controlMethodDict['100_method'] == 'Auto check':
        intervalText = '100%\n{}'.format(intervalText)

    if controlMethodDict['sample_no'] > 1:
        intervalText = '{}\nn=({})'.format(intervalText, controlMethodDict['sample_no'])
    return intervalText

def getControlMethodDetail(controlMethodDict: dict):
    if controlMethodDict.get('calibration_interval', '') != '':
        return 'Calibration'
    return ''

def getControlMethod(itemDict: dict):
    if (itemDict['control_method']['100_method'] == 'None'):
        return itemDict['control_item_type']
    return itemDict['control_method']['100_method']

def getMeasurement(itemDict: dict):
    finalText = itemDict['measurement']
    if itemDict['readability'] != '' or itemDict['parameter']['unit'] != '':
        finalText ='{} ({} {})'.format(
            finalText,
            itemDict['readability'],
            itemDict['parameter']['unit']
        )
    return finalText

def drawImage(img, row, col, rowOff, colOff):
    h, w = img.height, img.width
    size = XDRPositiveSize2D(p2e(w), p2e(h))
    marker = AnchorMarker(
        row=row,
        col=col,
        rowOff=p2e(rowOff),
        colOff=p2e(colOff)
    )
    img.anchor = OneCellAnchor(marker, size)
    return img

def getVerticalDashLine(height, row, col, rowOff, colOff):
    img = Image(drawVerticalDashedLine(EMU_to_pixels(c2e(height) * 0.45)))
    return drawImage(img, row, col, rowOff, colOff)

def getHorizontalDashLine(row, col, rowOff, colOff):
    return drawImage(
        Image('images/timing/dash-main-to-branch.png'),
        row,
        col,
        rowOff,
        colOff)

def getCheckProcess(row, col, rowOff, colOff):
    return drawImage(
        Image('images/timing/check-process.png'),
        row,
        col,
        rowOff,
        colOff)

def getCheckTimingSymbol(checkTiming, row, col, rowOff, colOff):
    symbolPath = checkTimingSymbolPathMap.get(checkTiming, None)
    if symbolPath is None:
        raise KeyError('Unregistered check timing type, {}'.format(checkTiming))
    img = Image(symbolPath)
    return drawImage(img, row, col, rowOff, colOff)

class PCSForm:
    def __init__(self, templatePath: str, dataDict: dict):
        self.templatePath = templatePath
        self.templateSheetName = 'empty'

        self.dataDict = dataDict
        self._initializeTemplateWorkbook()

    def _initializeTemplateWorkbook(self):
        self.workbook = load_workbook(filename = self.templatePath)

    def generate(self, fileName: str):
        headerDict = self.dataDict
        processList = self.dataDict['processes']

        templateSheet = self.workbook[self.templateSheetName]
        self._writeFormHeader(headerDict, templateSheet)

        totalProcess = len(processList)
        pageCount = 1
        for i, processDict in enumerate(processList):
            itemChunkList = chunk(processDict['items'], itemChunkSize)
            totalChunk = len(itemChunkList) if len(itemChunkList) > 1 else 1
            for j, itemChunk in enumerate(itemChunkList):
                itemSheet = self.workbook.copy_worksheet(templateSheet)
                self._writeFormProcess(
                    pageCount, totalProcess + totalChunk,
                    j+1, totalChunk,
                    processDict,
                    itemSheet)
                itemSheet.title = 'process-{}-{}'.format(
                    i+1,
                    j+1
                )
                self._writeProcessItem(
                    itemChunkSize * (j),
                    itemSheet,
                    itemChunk,
                    processDict['items']
                )
                pageCount = pageCount + 1
        
        self._saveWorkbook(fileName)

    def _writeFormHeader(self, headerDict: dict, sheet: Worksheet):
        #   Write check box
        sheet.cell(row=3, column=14).value = '❑    \t  Prototype'
        sheet.cell(row=4, column=14).value = '❑    \t  Pre-Launch'
        sheet.cell(row=5, column=14).value = '❑    \t  Production'

        sheet.cell(row=7, column=1).value = headerDict['line']
        sheet.cell(row=7, column=1).alignment = leftCenterAlignment
        sheet.cell(row=7, column=1).font = headerNormalStyle
        sheet.cell(row=7, column=8).value = headerDict['assy_name']
        sheet.cell(row=7, column=8).alignment = leftCenterAlignment
        sheet.cell(row=7, column=8).font = headerNormalStyle
        sheet.cell(row=9, column=8).value = headerDict['part_name']
        sheet.cell(row=9, column=8).alignment = leftCenterAlignment
        sheet.cell(row=9, column=8).font = headerNormalStyle
        sheet.cell(row=9, column=13).value = headerDict['customer']
        sheet.cell(row=9, column=13).alignment = centerCenterAlignment
        sheet.cell(row=9, column=13).font = headerNormalStyle

        sheet.cell(row=63, column=7).value = '                   Issue to ❑ Insp.    ❑ Prod.(___________)'

    def _writeFormProcess(self, idx: int, total: int, subIdx: int, subTotal: int, processDict: dict, sheet: Worksheet):
        #   Add denso logo
        densoIconImage = Image('images/denso-logo.png')
        h, w = densoIconImage.height, densoIconImage.width
        size = XDRPositiveSize2D(p2e(w), p2e(h))
        marker = AnchorMarker(
            row=0,
            col=0,
        )
        densoIconImage.anchor = OneCellAnchor(marker, size)
        sheet.add_image(densoIconImage)

        sheet.cell(row=2, column=15).value = 'Page  {} / {}'.format(
            idx,
            total
        )

        sheet.cell(row=9, column=1).value = '{}                  {} / {}'.format(
            processDict['name'],
            subIdx,
            subTotal
        )
        sheet.cell(row=9, column=1).alignment = leftCenterAlignment
        sheet.cell(row=9, column=1).font = headerNormalStyle

    def _writeProcessItem(self, startNumber:int, sheet: Worksheet, itemList: list, totalItemList: list):
        startRow = 12
        rowStep = 3
        startSeparatorColumn = 3
        endSeparatorColumn = 15
        
        isInheritGroup = False
        groupDashValue = None
        groupDashStart = None
        groupDashLength = None

        duringList = list()
        beforeList = list()
        afterList = list()

        for i, itemDict in enumerate(totalItemList):
            if itemDict['check_timing'] == 'During':
                duringList.append(i)
            elif itemDict['check_timing'] == 'Before':
                beforeList.append(i)
            elif itemDict['check_timing'] == 'After':
                afterList.append(i)

        #   Dash line
        # vertDashImg = getVerticalDashLine(len(itemList) * 3, 11, 0, -3, 8)
        # sheet.add_image(vertDashImg)

        landProcessIndex = None
        landProcessRow = None

        if len(duringList) > 0:
            if len(beforeList) == 0 and len(afterList) == 0:
                landIndex = len(totalItemList) / 2
                if landIndex < len(itemList) + startNumber - 1:
                    horizontalControlItemImg = getHorizontalDashLine(
                        startRow + (rowStep * (landIndex - startNumber)), 0, 7, 10
                    )
                    sheet.add_image(horizontalControlItemImg)
                    landProcessRow = startRow + (rowStep * (landIndex - startNumber))
                    landProcessIndex = int(landIndex)
            
            elif len(beforeList) == 0:
                if startNumber == 0:
                    horizontalControlItemImg = getHorizontalDashLine(
                        startRow, 0, 7, 10
                    )
                    sheet.add_image(horizontalControlItemImg)
                    landProcessRow = startRow
                    landProcessIndex = 0

            else:
                landIndex = duringList[int(len(duringList) / 2)]
                if landIndex < len(itemList) + startNumber - 1:
                    horizontalControlItemImg = getHorizontalDashLine(
                        startRow + (rowStep * (landIndex - startNumber)), 0, 7, 10
                    )
                    sheet.add_image(horizontalControlItemImg)
                    landProcessRow = startRow + (rowStep * (landIndex - startNumber))
                    landProcessIndex = int(landIndex)
        else:
            if len(beforeList) == 0:
                if startNumber == 0:
                    landProcessRow = startRow
                    landProcessIndex = 0
            elif len(afterList) == 0:
                if startNumber > 0:
                    landProcessRow = startRow + (rowStep * (len(itemList) - 1) + 1)
                    landProcessIndex = len(itemList) - 1
            else:
                lastBefore = beforeList[-1]
                firstAfterIndex = afterList[0]
                landIndex = (lastBefore + firstAfterIndex) / 2

                if landIndex < len(itemList) + startNumber - 1:
                    landProcessRow = startRow + (rowStep * (landIndex - startNumber))
                    landProcessIndex = int(landIndex)

        def groupNodeList(nodeList: list):
            groupDict = dict()
            i = 0
            for node in nodeList:
                if groupDict.get(i, None) is None:
                    groupDict[i] = [node]
                else:
                    if groupDict[i][-1] == node - 1:
                        groupDict[i].append(node)
                    else:
                        i += 1
                        groupDict[i] = [node]
            return list(groupDict.values())

        beforeGroup = groupNodeList(beforeList)
        duringGroup = groupNodeList(duringList)
        afterGroup = groupNodeList(afterList)

        if  (len(duringGroup) > 0 and duringGroup[-1][-1] > (startNumber + 1) * itemChunkSize) or \
            (len(beforeGroup) > 0 and beforeGroup[-1][-1] > (startNumber + 1) * itemChunkSize) or \
            (len(afterGroup) > 0 and afterGroup[-1][-1] > (startNumber + 1) * itemChunkSize):
            vertDashImg = getVerticalDashLine((itemChunkSize - landProcessIndex) * 2.88, startRow + (landProcessIndex * rowStep), 0, 8, 8)
            sheet.add_image(vertDashImg)

        for duringNodeList in duringGroup:
            drawLineLength = None
            drawLineRow = None
            drawConnectorRow = None
            vertRowOff = 8
            vertColOff = 8

            if len(duringNodeList) == 1:
                if (duringNodeList[0] < len(itemList) + startNumber):
                    drawConnectorRow = startRow + (rowStep * (duringNodeList[0] - startNumber))
                
                if landProcessIndex < startNumber:
                    if duringNodeList[0] >= startNumber:
                        adjustedValue = duringNodeList[0] - startNumber + 1
                        drawLineLength = ((adjustedValue - 1) * 2.99 if adjustedValue > 0 else 0) + 1.5
                        drawLineRow = startRow - 1
                        vertRowOff = 0

            if  drawLineLength is not None and \
                drawLineRow is not None:
                vertDashImg = getVerticalDashLine(drawLineLength, drawLineRow, 0, vertRowOff, vertColOff)
                sheet.add_image(vertDashImg)
            
            if drawConnectorRow is not None:
                horizontalControlItemImg = getHorizontalDashLine(
                    drawConnectorRow, 0, 7, 10
                )
                sheet.add_image(horizontalControlItemImg)

        for afterNodeList in afterGroup:
            drawLineLength = None
            drawLineRow = None
            drawConnectorRow = None
            vertRowOff = 8
            vertColOff = 8

            afterCenterValue = afterNodeList[0] if len(afterNodeList) == 1 else afterNodeList[int(len(afterNodeList) / 2)]
            if landProcessIndex >= startNumber:
                drawLineLength = (afterCenterValue - landProcessIndex) * 2.99
                drawLineRow = startRow + (rowStep * landProcessIndex)
                drawConnectorRow = startRow + (rowStep * afterCenterValue)
            elif afterNodeList[-1] >= startNumber:
                adjustedCenterValue = ((afterCenterValue if afterCenterValue >= startNumber else startNumber) - startNumber)
                drawLineLength = ((adjustedCenterValue - 1) * 2.99 if adjustedCenterValue > 0 else 0) + 1.5
                drawLineRow = startRow - 1
                drawConnectorRow = startRow + (rowStep * adjustedCenterValue)
                vertRowOff = 0

            if  drawLineLength is not None and \
                drawLineRow is not None:
                vertDashImg = getVerticalDashLine(drawLineLength, drawLineRow, 0, vertRowOff, vertColOff)
                sheet.add_image(vertDashImg)
            
            if drawConnectorRow is not None:
                horizontalControlItemImg = getHorizontalDashLine(
                    drawConnectorRow, 0, 7, 10
                )
                sheet.add_image(horizontalControlItemImg)

        for beforeNodeList in beforeGroup:
            drawLineLength = None
            drawLineRow = None
            drawConnectorRow = None

            beforeCenterValue = beforeNodeList[0] if len(beforeNodeList) == 1 else beforeNodeList[int(len(beforeNodeList) / 2)]
            if landProcessIndex >= startNumber:
                drawLineLength = abs(beforeCenterValue - landProcessIndex) * 2.99
                drawLineRow = startRow + (rowStep * beforeCenterValue)
                drawConnectorRow = startRow + (rowStep * beforeCenterValue)
            elif beforeNodeList[-1] >= startNumber:
                adjustedCenterValue = ((beforeCenterValue if beforeCenterValue >= startNumber else startNumber) - startNumber)
                drawLineLength = ((adjustedCenterValue - 1) * 2.99 if adjustedCenterValue > 0 else 0) + 1.5
                drawLineRow = startRow
                drawConnectorRow = startRow + (rowStep * beforeCenterValue)

            if  drawLineLength is not None and \
                drawLineRow is not None:
                vertDashImg = getVerticalDashLine(drawLineLength, drawLineRow, 0, 8, 8)
                sheet.add_image(vertDashImg)
            if drawConnectorRow is not None:
                horizontalControlItemImg = getHorizontalDashLine(
                    drawConnectorRow, 0, 7, 10
                )
                sheet.add_image(horizontalControlItemImg)

        checkProcessImage = getCheckProcess(
            landProcessRow, 0, 0, 0
        )
        sheet.add_image(checkProcessImage)

        for i, item in enumerate(itemList):
            #   Cell merging
            sheet.merge_cells('E{}:H{}'.format(startRow + (rowStep * i), startRow + (rowStep * i)+1))
            sheet.merge_cells('E{}:H{}'.format(startRow + (rowStep * i) + 2, startRow + (rowStep * i) + 2))
            sheet.merge_cells('I{}:I{}'.format(startRow + (rowStep * i), startRow + (rowStep * i)+1))
            sheet.merge_cells('J{}:J{}'.format(startRow + (rowStep * i), startRow + (rowStep * i)+1))
            sheet.merge_cells('K{}:K{}'.format(startRow + (rowStep * i), startRow + (rowStep * i)+1))
            sheet.merge_cells('L{}:L{}'.format(startRow + (rowStep * i), startRow + (rowStep * i)+2))
            sheet.merge_cells('M{}:N{}'.format(startRow + (rowStep * i), startRow + (rowStep * i) + 2))
            sheet.merge_cells('O{}:O{}'.format(startRow + (rowStep * i), startRow + (rowStep * i) + 2))
            sheet.merge_cells('M7:O7')
            sheet.merge_cells('A9:G9')

            #   Cell bordering
            sheet.cell(row=startRow + (rowStep * i) + 1, column=5).border = bottomBorder
            sheet.cell(row=startRow + (rowStep * i) + 1, column=9).border = bottomBorder
            sheet.cell(row=startRow + (rowStep * i) + 1, column=10).border = bottomBorder
            sheet.cell(row=startRow + (rowStep * i) + 1, column=11).border = bottomBorder
            for j in range(endSeparatorColumn - startSeparatorColumn):
                sheet.cell(row=startRow + (rowStep * i) + 2, column=startSeparatorColumn + j + 1).border = bottomRightBorder

            #   Cell values
            sheet.cell(row=startRow + (rowStep * i), column=4).value = startNumber + (i + 1)
            sheet.cell(row=startRow + (rowStep * i), column=4).alignment = centerCenterAlignment
            sheet.cell(row=startRow + (rowStep * i), column=4).font = textNormalStyle
            sheet.cell(row=(startRow + (rowStep * i)), column=5).value = getParameter(item['parameter'])
            sheet.cell(row=(startRow + (rowStep * i)), column=5).font = textNormalStyle
            sheet.cell(row=(startRow + (rowStep * i) + 2), column=5).value = getMeasurement(item)
            sheet.cell(row=(startRow + (rowStep * i) + 2), column=5).font = textNormalStyle
            sheet.cell(row=(startRow + (rowStep * i)), column=5).alignment = topLeftAlignment
            sheet.cell(row=(startRow + (rowStep * i)), column=9).value = getInterval(item['control_method'])
            sheet.cell(row=(startRow + (rowStep * i)), column=9).font = textNormalStyle
            sheet.cell(row=(startRow + (rowStep * i)), column=9).alignment = topCenterAlignment
            sheet.cell(row=(startRow + (rowStep * i) + 2), column=9).value = item['control_method'].get('calibration_interval', '')
            sheet.cell(row=(startRow + (rowStep * i) + 2), column=9).alignment = centerCenterAlignment
            sheet.cell(row=(startRow + (rowStep * i) + 2), column=9).font = textNormalStyle
            sheet.cell(row=(startRow + (rowStep * i)), column=10).value = getControlMethod(item)
            sheet.cell(row=(startRow + (rowStep * i)), column=10).alignment = centerCenterAlignment
            sheet.cell(row=(startRow + (rowStep * i)), column=10).font = textNormalStyle
            sheet.cell(row=(startRow + (rowStep * i)) + 2, column=10).value = getControlMethodDetail(item['control_method'])
            sheet.cell(row=(startRow + (rowStep * i)) + 2, column=10).alignment = centerCenterAlignment
            sheet.cell(row=(startRow + (rowStep * i)) + 2, column=10).font = textNormalStyle
            sheet.cell(row=(startRow + (rowStep * i)), column=11).value = item['control_method']['in_charge']
            sheet.cell(row=(startRow + (rowStep * i)), column=11).alignment = centerCenterAlignment
            sheet.cell(row=(startRow + (rowStep * i)), column=11).font = textNormalStyle
            sheet.cell(row=(startRow + (rowStep * i)), column=12).value = getProcessCapability(item['initial_p_capability'])
            sheet.cell(row=(startRow + (rowStep * i)), column=12).font = textNormalStyle
            sheet.cell(row=(startRow + (rowStep * i)), column=13).value = item['remark']['remark']
            sheet.cell(row=(startRow + (rowStep * i)), column=13).alignment = topLeftAlignment
            sheet.cell(row=(startRow + (rowStep * i)), column=13).font = textNormalStyle
            sheet.cell(row=(startRow + (rowStep * i)), column=15).value = item['remark']['ws_no']
            sheet.cell(row=(startRow + (rowStep * i)), column=15).alignment = centerCenterAlignment
            sheet.cell(row=(startRow + (rowStep * i)), column=15).font = textNormalStyle

            #   Imaging

            #   Grouping
            currentIndexInTotal = i + startNumber
            if i == 0 and startNumber != 0 and totalItemList[currentIndexInTotal - 1]['check_timing'] == item['check_timing']:
                groupDashValue = item['check_timing']
                groupDashStart = 0
                groupDashLength = 0
                isInheritGroup = True

            if groupDashValue == item['check_timing']:
                groupDashLength += 1

            if groupDashValue is None:
                groupDashValue = item['check_timing']
                groupDashStart = i
                groupDashLength = 1
            
            if currentIndexInTotal + 1 >= len(totalItemList):
                if groupDashLength > 1:
                    if isInheritGroup:
                        print('in')
                        vertDashImg = getVerticalDashLine((groupDashLength * 2.99) - 3 + 1.5, startRow + (rowStep * groupDashStart), 1, 8 - 28, 4)
                    else:
                        vertDashImg = getVerticalDashLine((groupDashLength * 2.99) - 3, startRow + (rowStep * groupDashStart), 1, 8, 4)
                    sheet.add_image(vertDashImg)
                groupDashValue = None
                groupDashStart = None
                groupDashLength = None
                isInheritGroup = False

            elif i + 1 >= len(itemList):
                if groupDashLength > 1:
                    if isInheritGroup:
                        print('in')
                        vertDashImg = getVerticalDashLine((groupDashLength * 2.99) - 3 + 1.5 + 1.5 if currentIndexInTotal + 1 < len(totalItemList) and groupDashValue == totalItemList[currentIndexInTotal + 1]['check_timing'] else 0, startRow + (rowStep * groupDashStart), 1, 8 - 28, 4)
                    else:
                        vertDashImg = getVerticalDashLine((groupDashLength * 2.99) - 3 + 1.5 if currentIndexInTotal + 1 < len(totalItemList) and groupDashValue == totalItemList[currentIndexInTotal + 1]['check_timing'] else 0, startRow + (rowStep * groupDashStart), 1, 8, 4)

                    sheet.add_image(vertDashImg)
                groupDashValue = None
                groupDashStart = None
                groupDashLength = None
                isInheritGroup = False

            elif i + 1 < len(itemList):
                if itemList[i + 1]['check_timing'] != groupDashValue:
                    if groupDashLength > 1:
                        if isInheritGroup:
                            print('in')
                            vertDashImg = getVerticalDashLine((groupDashLength * 2.99) - 3 + 1.5, startRow + (rowStep * groupDashStart), 1, 8 - 28, 4)
                        else:
                            vertDashImg = getVerticalDashLine((groupDashLength * 2.99) - 3, startRow + (rowStep * groupDashStart), 1, 8, 4)
                        sheet.add_image(vertDashImg)
                    groupDashValue = None
                    groupDashStart = None
                    groupDashLength = None
                    isInheritGroup = False

            scSymbolImgList = getSCSymbolList(item['sc_symbols'], startRow + (rowStep * i), 3)
            for scSymbol in scSymbolImgList:
                sheet.add_image(scSymbol)

            horizontalControlItemImg = getHorizontalDashLine(
            startRow + (rowStep * i), 1, 7.5, 3
            )
            sheet.add_image(horizontalControlItemImg)

            totalScSymbolList = getTotalSCSymbolList(itemList)
            for totalScSymbol in totalScSymbolList:
                sheet.add_image(totalScSymbol)

            controlItemSymbolImg = getCheckTimingSymbol(
                item['control_item_type'],
                startRow + (rowStep * i), 1, 0, 10)
            sheet.add_image(controlItemSymbolImg)

    def _saveWorkbook(self, fileName: str):
        templateSheet = self.workbook[self.templateSheetName]
        self.workbook.remove(templateSheet)
        self.workbook.save(getOutputFilePath(fileName))
        
